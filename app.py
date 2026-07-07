"""Email Pitch Tool - 轻量级邮件营销工具 MVP"""
import os
import io
import csv
import json
import re
import time  # ✏️ CHANGED: added for cache-busting timestamps
import threading
import sqlite3
import base64
import html as html_module
import hashlib as _hashlib
import hmac as _hmac
from datetime import datetime, timedelta
from pathlib import Path
from contextlib import contextmanager
from typing import Optional
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from jinja2 import Template
from openpyxl import load_workbook
from email_validator import validate_email, EmailNotValidError
import secrets as _secrets
from fastapi import FastAPI, Depends, UploadFile, File, Form, HTTPException, Request
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, Response  # ✏️ CHANGED: added Response for tracking GIF
from fastapi.staticfiles import StaticFiles
from apscheduler.schedulers.background import BackgroundScheduler
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build

def build_gmail_service(creds):
    try:
        import httplib2
        from google_auth_httplib2 import AuthorizedHttp
        http = AuthorizedHttp(creds, http=httplib2.Http(timeout=20))
        return build('gmail', 'v1', http=http, cache_discovery=False)
    except Exception as e:
        print(f"[Gmail client warning] falling back without explicit timeout: {e}")
        return build('gmail', 'v1', credentials=creds, cache_discovery=False)

app = FastAPI(title="Email Pitch Tool")
scheduler = BackgroundScheduler()
scheduler.start()

# Per-campaign lock prevents duplicate sends when Launch is clicked twice or overlaps scheduler.
campaign_process_locks = {}
campaign_process_locks_guard = threading.Lock()

# HTTP Basic Auth
_security = HTTPBasic()
_AUTH_USER = os.environ.get("AUTH_USERNAME", "admin")
_AUTH_PASS = os.environ.get("AUTH_PASSWORD", "admin")
_AUTH_COOKIE = "email_pitch_auth"

# Paths that skip auth (tracking pixels must work without login)
_PUBLIC_PATHS = {"/track/open", "/track/click", "/oauth/callback", "/static"}


def _auth_cookie_value() -> str:
    key = (_AUTH_PASS or "admin").encode("utf-8")
    msg = f"{_AUTH_USER}:email-pitch-tool".encode("utf-8")
    return _hmac.new(key, msg, _hashlib.sha256).hexdigest()


def _has_valid_auth_cookie(request: Request) -> bool:
    cookie = request.cookies.get(_AUTH_COOKIE, "")
    return bool(cookie) and _secrets.compare_digest(cookie, _auth_cookie_value())


def _set_auth_cookie(response, request: Request):
    forwarded_proto = request.headers.get("x-forwarded-proto", "").lower()
    secure = request.url.scheme == "https" or forwarded_proto == "https"
    response.set_cookie(
        _AUTH_COOKIE,
        _auth_cookie_value(),
        max_age=7 * 24 * 60 * 60,
        httponly=True,
        secure=secure,
        samesite="lax",
    )


@app.middleware("http")
async def basic_auth_middleware(request: Request, call_next):
    path = request.url.path.rstrip("/")
    # Skip auth for tracking & OAuth callback
    for pub in _PUBLIC_PATHS:
        if path.startswith(pub):
            return await call_next(request)

    # Browser fetch does not always resend cached Basic credentials. Accept a
    # cookie set after the first successful Basic Auth challenge so preview/API
    # calls do not repeatedly trigger login dialogs.
    if _has_valid_auth_cookie(request):
        return await call_next(request)

    # Check Basic Auth header
    auth = request.headers.get("authorization", "")
    if auth.startswith("Basic "):
        import base64 as _b64
        try:
            decoded = _b64.b64decode(auth[6:]).decode()
            user, pwd = decoded.split(":", 1)
            if _secrets.compare_digest(user, _AUTH_USER) and _secrets.compare_digest(pwd, _AUTH_PASS):
                response = await call_next(request)
                _set_auth_cookie(response, request)
                return response
        except Exception:
            pass
    from starlette.responses import Response as _Resp
    return _Resp(
        status_code=401,
        headers={"WWW-Authenticate": 'Basic realm="Email Pitch Tool"'},
        content="Unauthorized"
    )

# 配置
DB_PATH = "data.db"
CREDENTIALS_FILE = "credentials.json"  # 从Google Cloud Console下载
SCOPES = ['https://www.googleapis.com/auth/gmail.modify']  # covers send + read + labels + thread modify
# 支持环境变量配置REDIRECT_URI（部署时使用）
BASE_URL = os.environ.get("BASE_URL", "http://localhost:8000")
REDIRECT_URI = f"{BASE_URL}/oauth/callback"
TEST_MODE = os.environ.get("TEST_MODE", "false").lower() == "true"  # 测试模式不发真邮件

# 数据库初始化
def init_db():
    conn = sqlite3.connect(DB_PATH, timeout=30.0)
    # 启用WAL模式以提高并发性能
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")  # 30秒超时
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS accounts (
            id INTEGER PRIMARY KEY, email TEXT UNIQUE, token TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS campaigns (
            id INTEGER PRIMARY KEY, name TEXT, status TEXT DEFAULT 'draft',
            account_email TEXT, interval_minutes INTEGER DEFAULT 5,
            cc_email TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS templates (
            id INTEGER PRIMARY KEY, campaign_id INTEGER, step INTEGER, subject TEXT, body TEXT,
            delay_days INTEGER DEFAULT 0,
            delay_hours INTEGER DEFAULT 0,
            FOREIGN KEY(campaign_id) REFERENCES campaigns(id)
        );
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY, campaign_id INTEGER, email TEXT, data TEXT,
            status TEXT DEFAULT 'pending', current_step INTEGER DEFAULT 1,
            last_sent_at TIMESTAMP, opened INTEGER DEFAULT 0, clicked INTEGER DEFAULT 0, replied INTEGER DEFAULT 0,
            FOREIGN KEY(campaign_id) REFERENCES campaigns(id)
        );
        CREATE TABLE IF NOT EXISTS blacklist (email TEXT PRIMARY KEY);

        CREATE TABLE IF NOT EXISTS oauth_states (
            state TEXT PRIMARY KEY,
            code_verifier TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        /* ✏️ CHANGED: new settings table to store BASE_URL from the web UI */
        CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT);
    """)

    # ✏️ CC support: migrate existing databases
    try:
        conn.execute("ALTER TABLE campaigns ADD COLUMN cc_email TEXT DEFAULT ''")
    except:
        pass  # column already exists

    # ✏️ Threading support: store message_id and thread_id for follow-up threading
    for col in ('message_id', 'thread_id'):
        try:
            conn.execute(f"ALTER TABLE leads ADD COLUMN {col} TEXT DEFAULT ''")
        except:
            pass  # column already exists

    # ✏️ Follow-up delay support: allow hour-level delays for faster tests
    try:
        conn.execute("ALTER TABLE templates ADD COLUMN delay_hours INTEGER DEFAULT 0")
    except:
        pass  # column already exists
    conn.commit()
    conn.close()
init_db()

# ✏️ CHANGED: helper to get the tracking base URL (checks DB settings first, then env var, then localhost)
def get_tracking_base_url():
    """Get BASE_URL for tracking pixels: DB setting > env var > localhost fallback"""
    try:
        with get_db() as conn:
            row = conn.execute("SELECT value FROM settings WHERE key='base_url'").fetchone()
            if row and row['value']:
                return row['value'].rstrip('/')
    except Exception:
        pass
    return os.environ.get("BASE_URL", "http://localhost:8000").rstrip('/')

# Markdown-to-HTML email body processor with list support
def prepare_email_body(body: str) -> str:
    """If body has no HTML tags, treat as plain text with Markdown support.
    Supports: ![alt](url), [text](url), **bold**, *italic*, bare URLs, lists, newlines."""
    if re.search(r'<[a-zA-Z][^>]*>', body):
        return body  # already HTML

    # Normalize line endings (browser forms send \r\n)
    body = body.replace('\r\n', '\n').replace('\r', '\n')

    # Process Markdown syntax BEFORE escaping (so URLs stay intact)
    body = re.sub(r'!\[([^\]]*)\]\(([^)]+)\)', r'<img src="\2" alt="\1" style="max-width:100%">', body)
    body = re.sub(r'\[([^\]]+)\]\(([^)]+)\)', r'<a href="\2">\1</a>', body)
    body = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', body)
    body = re.sub(r'(?<!\*)\*(?!\*)(.+?)(?<!\*)\*(?!\*)', r'<i>\1</i>', body)

    # Process Markdown lists: convert '- item' to <ul><li> and '1. item' to <ol><li>
    lines = body.split('\n')
    processed = []
    in_list = False
    list_type = None
    for line in lines:
        stripped = line.strip()
        is_ul = bool(re.match(r'^[-]\s+', stripped))
        is_ol = bool(re.match(r'^\d+\.\s+', stripped))
        if is_ul:
            item_text = re.sub(r'^[-]\s+', '', stripped)
            if not in_list or list_type != 'ul':
                if in_list:
                    processed.append('</' + list_type + '>')
                processed.append('<ul style="margin:8px 0;padding-left:20px">')
                in_list = True
                list_type = 'ul'
            processed.append('<li style="margin:2px 0">' + item_text + '</li>')
        elif is_ol:
            item_text = re.sub(r'^\d+\.\s+', '', stripped)
            if not in_list or list_type != 'ol':
                if in_list:
                    processed.append('</' + list_type + '>')
                processed.append('<ol style="margin:8px 0;padding-left:20px">')
                in_list = True
                list_type = 'ol'
            processed.append('<li style="margin:2px 0">' + item_text + '</li>')
        else:
            if in_list:
                processed.append('</' + list_type + '>')
                in_list = False
                list_type = None
            processed.append(line)
    if in_list:
        processed.append('</' + list_type + '>')
    body = '\n'.join(processed)

    # Escape plain text parts only (not the HTML tags we just created)
    parts = re.split(r'(<[^>]+>)', body)
    result = []
    for part in parts:
        if part.startswith('<') and part.endswith('>'):
            result.append(part)  # keep HTML tags as-is
        else:
            escaped = html_module.escape(part)
            # Auto-link bare URLs in text
            escaped = re.sub(r'(https?://[^\s<]+)', r'<a href="\1">\1</a>', escaped)
            result.append(escaped)

    body = ''.join(result)
    return body.replace('\n', '<br>\n')


def normalize_template_key(key: str) -> str:
    """Normalize CSV headers and template vars for forgiving placeholder matching."""
    return re.sub(r'[^a-z0-9]+', '_', str(key or '').strip().lower()).strip('_')


def lookup_template_value(key: str, data: dict):
    if key in data:
        return data.get(key)
    wanted = normalize_template_key(key)
    if not wanted:
        return None
    for existing_key, value in data.items():
        if normalize_template_key(existing_key) == wanted:
            return value
    return None


def render_template_text(template_text: str, data: dict) -> str:
    """Render placeholders while supporting CSV headers like {{Company Name}}."""
    def replace_raw_placeholder(match):
        key = match.group(1).strip()
        default = match.group(2).strip() if match.group(2) is not None else None
        value = lookup_template_value(key, data)
        if value is not None and str(value) != "":
            return str(value)
        if default is not None:
            return default
        # Keep normal Jinja identifiers available for the legacy renderer.
        if re.match(r'^[A-Za-z_][A-Za-z0-9_\.]*$', key):
            return match.group(0)
        return ""

    normalized = re.sub(r'\{\{\s*([^{}|]+?)\s*(?:\|([^{}]*))?\}\}', replace_raw_placeholder, template_text or "")
    return Template(normalized).render(**data)

def threaded_reply_subject(original_subject: str) -> str:
    """Use a stable reply subject so Gmail keeps follow-ups in the same thread."""
    subject = (original_subject or '').strip()
    if subject.lower().startswith('re:'):
        return subject
    return f"Re: {subject}" if subject else subject

def get_followup_subject(conn, cid: int, data: dict, fallback_subject: str = "") -> str:
    """Follow-up emails always reuse Step 1's subject for Gmail threading."""
    first_tpl = conn.execute(
        "SELECT subject FROM templates WHERE campaign_id=? AND step=1 ORDER BY id LIMIT 1",
        (cid,)
    ).fetchone()
    base_subject = fallback_subject
    if first_tpl and first_tpl['subject']:
        base_subject = render_template_text(first_tpl['subject'], data)
    return threaded_reply_subject(base_subject)

def check_all_replies():
    """定期检查所有campaign的回复"""
    try:
        with get_db() as conn:
            campaigns = conn.execute(
                "SELECT DISTINCT c.id, c.account_email FROM campaigns c "
                "JOIN leads l ON c.id = l.campaign_id "
                "WHERE c.account_email IS NOT NULL AND l.replied = 0"
            ).fetchall()

        # 释放数据库连接后再逐个检查
        for campaign in campaigns:
            try:
                check_replies(campaign['id'], campaign['account_email'])
            except Exception as e:
                print(f"[Check replies error for campaign {campaign['id']}] {e}")
    except Exception as e:
        print(f"[Check all replies error] {e}")

def restore_running_campaigns():
    """启动时恢复所有运行中的campaigns"""
    with get_db() as conn:
        rows = conn.execute("SELECT id, account_email, interval_minutes FROM campaigns WHERE status='running'").fetchall()

    for row in rows:
        if row['account_email']:
            scheduler.add_job(process_campaign, 'interval', minutes=row['interval_minutes'] or 5,
                              args=[row['id'], row['account_email']], id=f"campaign_{row['id']}", replace_existing=True)
            print(f"[Restored] Campaign {row['id']} with interval {row['interval_minutes']}min")

    # 添加定期回复检查任务（每10分钟检查一次所有campaign）
    scheduler.add_job(check_all_replies, 'interval', minutes=10, id='check_all_replies', replace_existing=True)
    print("[Scheduled] Reply checker every 10 minutes")

@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30.0)
    conn.execute("PRAGMA busy_timeout=30000")  # 30秒超时
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()

# Gmail OAuth
@app.get("/oauth/start")
def oauth_start():
    # 支持环境变量配置（用于云端部署）
    if os.environ.get("GOOGLE_CLIENT_ID") and os.environ.get("GOOGLE_CLIENT_SECRET"):
        client_config = {
            "web": {
                "client_id": os.environ.get("GOOGLE_CLIENT_ID"),
                "client_secret": os.environ.get("GOOGLE_CLIENT_SECRET"),
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [REDIRECT_URI]
            }
        }
        flow = Flow.from_client_config(client_config, scopes=SCOPES, redirect_uri=REDIRECT_URI)
    elif Path(CREDENTIALS_FILE).exists():
        flow = Flow.from_client_secrets_file(CREDENTIALS_FILE, scopes=SCOPES, redirect_uri=REDIRECT_URI)
    else:
        return {"error": "请先配置 Google OAuth 凭据（credentials.json 或环境变量）"}

    auth_url, state = flow.authorization_url(prompt='consent')
    code_verifier = getattr(flow, "code_verifier", None)
    if not code_verifier:
        code_verifier = getattr(flow.oauth2session._client, "code_verifier", None)
    if code_verifier:
        with get_db() as conn:
            conn.execute("DELETE FROM oauth_states WHERE created_at < datetime('now', '-1 hour')")
            conn.execute(
                "INSERT OR REPLACE INTO oauth_states(state, code_verifier) VALUES(?, ?)",
                (state, code_verifier),
            )
            conn.commit()
    return RedirectResponse(auth_url)

@app.get("/oauth/callback")
def oauth_callback(code: str, state: str = ""):
    # 使用动态REDIRECT_URI
    redirect_uri = f"{BASE_URL}/oauth/callback"

    # 支持环境变量配置
    if os.environ.get("GOOGLE_CLIENT_ID") and os.environ.get("GOOGLE_CLIENT_SECRET"):
        client_config = {
            "web": {
                "client_id": os.environ.get("GOOGLE_CLIENT_ID"),
                "client_secret": os.environ.get("GOOGLE_CLIENT_SECRET"),
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [redirect_uri]
            }
        }
        flow = Flow.from_client_config(client_config, scopes=SCOPES, redirect_uri=redirect_uri)
    else:
        flow = Flow.from_client_secrets_file(CREDENTIALS_FILE, scopes=SCOPES, redirect_uri=redirect_uri)

    code_verifier = None
    if state:
        with get_db() as conn:
            row = conn.execute("SELECT code_verifier FROM oauth_states WHERE state=?", (state,)).fetchone()
            if row:
                code_verifier = row["code_verifier"]
                conn.execute("DELETE FROM oauth_states WHERE state=?", (state,))
                conn.commit()

    if code_verifier:
        flow.fetch_token(code=code, code_verifier=code_verifier)
    else:
        flow.fetch_token(code=code)
    creds = flow.credentials
    service = build_gmail_service(creds)
    profile = service.users().getProfile(userId='me').execute()
    email = profile['emailAddress']
    with get_db() as conn:
        conn.execute("INSERT OR REPLACE INTO accounts(email, token) VALUES(?, ?)",
                    (email, creds.to_json()))
        conn.commit()
    return RedirectResponse("/?msg=账号绑定成功: " + email)

# API 路由
@app.post("/api/campaigns")
def create_campaign(name: str = Form(...)):
    with get_db() as conn:
        cur = conn.execute("INSERT INTO campaigns(name) VALUES(?)", (name,))
        conn.commit()
        return {"id": cur.lastrowid, "name": name}

@app.get("/api/campaigns")
def list_campaigns():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM campaigns ORDER BY id DESC").fetchall()
        return [dict(r) for r in rows]

@app.post("/api/campaigns/{cid}/templates")
def add_template(cid: int, step: int = Form(...), subject: str = Form(""), body: str = Form(...), delay_days: int = Form(0), delay_hours: int = Form(0)):
    with get_db() as conn:
        conn.execute("INSERT INTO templates(campaign_id, step, subject, body, delay_days, delay_hours) VALUES(?,?,?,?,?,?)",
                    (cid, step, subject, body, delay_days, delay_hours))
        conn.commit()
    return {"ok": True}

@app.get("/api/campaigns/{cid}/templates")
def get_templates(cid: int):
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM templates WHERE campaign_id=? ORDER BY step", (cid,)).fetchall()
        return [dict(r) for r in rows]

@app.put("/api/templates/{tid}")
def update_template(tid: int, step: int = Form(...), subject: str = Form(""), body: str = Form(...), delay_days: int = Form(0), delay_hours: int = Form(0)):
    with get_db() as conn:
        conn.execute("UPDATE templates SET step=?, subject=?, body=?, delay_days=?, delay_hours=? WHERE id=?",
                    (step, subject, body, delay_days, delay_hours, tid))
        conn.commit()
    return {"ok": True}

@app.delete("/api/templates/{tid}")
def delete_template(tid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM templates WHERE id=?", (tid,))
        conn.commit()
    return {"ok": True}

@app.get("/api/campaigns/{cid}/variables")
def get_campaign_variables(cid: int):
    """提取campaign所有模板中使用的变量"""
    import re
    variables = set()
    with get_db() as conn:
        rows = conn.execute("SELECT subject, body FROM templates WHERE campaign_id=?", (cid,)).fetchall()
        for row in rows:
            # 匹配 {{var}} 或 {{var|default}}
            for text in [row['subject'], row['body']]:
                if text:
                    matches = re.findall(r'\{\{\s*([^{}|]+?)(?:\|[^}]*)?\s*\}\}', text)
                    variables.update(m.strip() for m in matches if m.strip())
    # email是必须的，不需要提示
    variables.discard('email')
    return {"variables": sorted(variables)}

@app.post("/api/campaigns/{cid}/leads")
async def upload_leads(cid: int, file: UploadFile = File(...), defaults: str = Form("{}")):
    content = await file.read()
    default_values = json.loads(defaults)
    rows = []

    if file.filename.endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    else:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

    if not rows or 'email' not in rows[0]:
        raise HTTPException(400, "文件必须包含email列")

    added, skipped = 0, 0
    with get_db() as conn:
        blacklist = {r[0] for r in conn.execute("SELECT email FROM blacklist").fetchall()}
        existing = {r[0] for r in conn.execute("SELECT email FROM leads WHERE campaign_id=?", (cid,)).fetchall()}

        for row in rows:
            email = str(row.get('email', '')).strip().lower()
            try:
                validate_email(email)
                if email in blacklist or email in existing:
                    skipped += 1; continue
                # 合并默认值和行数据
                data = {**default_values, **{k: v for k, v in row.items() if k != 'email'}}
                conn.execute("INSERT INTO leads(campaign_id, email, data) VALUES(?,?,?)",
                            (cid, email, json.dumps(data, default=str)))
                added += 1
            except EmailNotValidError:
                skipped += 1
        conn.commit()
    return {"added": added, "skipped": skipped}

@app.get("/api/campaigns/{cid}/leads")
def get_leads(cid: int):
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM leads WHERE campaign_id=?", (cid,)).fetchall()
        return [dict(r) for r in rows]

@app.post("/api/campaigns/{cid}/leads/json")
async def add_leads_json(cid: int, request: Request):
    """支持JSON格式批量导入leads"""
    rows = await request.json()

    added, skipped = 0, 0
    errors = []
    with get_db() as conn:
        blacklist = {r[0] for r in conn.execute("SELECT email FROM blacklist").fetchall()}
        existing = {r[0] for r in conn.execute("SELECT email FROM leads WHERE campaign_id=?", (cid,)).fetchall()}

        for row in rows:
            email = str(row.get('email', '')).strip().lower()
            if not email:
                errors.append("空邮箱")
                skipped += 1
                continue
            try:
                # 使用宽松模式验证邮箱
                validate_email(email, check_deliverability=False)
                if email in blacklist:
                    errors.append(f"{email}: 在黑名单中")
                    skipped += 1
                    continue
                if email in existing:
                    errors.append(f"{email}: 已存在")
                    skipped += 1
                    continue
                data = {k: v for k, v in row.items() if k != 'email'}
                conn.execute("INSERT INTO leads(campaign_id, email, data) VALUES(?,?,?)",
                            (cid, email, json.dumps(data, default=str)))
                existing.add(email)
                added += 1
            except EmailNotValidError as e:
                errors.append(f"{email}: {str(e)}")
                skipped += 1
        conn.commit()
    return {"added": added, "skipped": skipped, "errors": errors[:5]}

@app.get("/api/campaigns/{cid}/preview")
def preview_email(cid: int, lead_id: int, step: int = 1):
    with get_db() as conn:
        lead = conn.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
        tpl = conn.execute("SELECT * FROM templates WHERE campaign_id=? AND step=?", (cid, step)).fetchone()
        if not lead or not tpl: raise HTTPException(404)
        data = json.loads(lead['data'])
        data['email'] = lead['email']
        subject = render_template_text(tpl['subject'], data)
        if step > 1:
            subject = get_followup_subject(conn, cid, data, subject)
        return {
            "subject": subject,
            "body": prepare_email_body(render_template_text(tpl['body'], data))
        }

@app.post("/api/campaigns/{cid}/save-settings")
def save_campaign_settings(cid: int, account_email: str = Form(""), interval_minutes: int = Form(5), cc_email: str = Form("")):
    with get_db() as conn:
        conn.execute("UPDATE campaigns SET account_email=?, interval_minutes=?, cc_email=? WHERE id=?",
                    (account_email, interval_minutes, cc_email, cid))
        conn.commit()
    return {"ok": True, "msg": "设置已保存"}

@app.post("/api/campaigns/{cid}/launch")
def launch_campaign(cid: int, account_email: str = Form(...), interval_minutes: int = Form(5), cc_email: str = Form('')):
    lock = _get_campaign_process_lock(cid)
    if not lock.acquire(timeout=60):
        raise HTTPException(status_code=409, detail="Campaign 正在启动中，请稍后再试")

    try:
        with get_db() as conn:
            row = conn.execute("SELECT status FROM campaigns WHERE id=?", (cid,)).fetchone()
            if not row:
                raise HTTPException(404)
            was_running = row['status'] == 'running'
            conn.execute("UPDATE campaigns SET status='running', account_email=?, interval_minutes=?, cc_email=? WHERE id=?",
                        (account_email, interval_minutes, cc_email, cid))
            conn.commit()

        # 添加/刷新定时任务继续发送后续邮件
        scheduler.add_job(process_campaign, 'interval', minutes=interval_minutes,
                          args=[cid, account_email], id=f"campaign_{cid}", replace_existing=True)

        if was_running:
            return {"ok": True, "msg": f"Campaign 已在运行，已更新为每{interval_minutes}分钟检查一次"}

        # 首次启动时立即处理一封；Launch 本身也串行化，避免双击触发两次即时发送。
        _process_campaign_locked(cid, account_email)
        return {"ok": True, "msg": f"已启动，立即发送第一封，之后每{interval_minutes}分钟检查一次"}
    finally:
        lock.release()

@app.get("/api/campaigns/{cid}")
def get_campaign(cid: int):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM campaigns WHERE id=?", (cid,)).fetchone()
        if not row: raise HTTPException(404)
        return dict(row)

@app.post("/api/campaigns/{cid}/stop")
def stop_campaign(cid: int):
    with get_db() as conn:
        conn.execute("UPDATE campaigns SET status='paused' WHERE id=?", (cid,))
        conn.commit()
    try: scheduler.remove_job(f"campaign_{cid}")
    except: pass
    return {"ok": True}

@app.delete("/api/campaigns/{cid}")
def delete_campaign(cid: int):
    """删除 campaign 及其所有相关数据"""
    # 停止 scheduler job
    try:
        scheduler.remove_job(f"campaign_{cid}")
    except:
        pass

    # 删除数据库记录
    with get_db() as conn:
        conn.execute("DELETE FROM leads WHERE campaign_id=?", (cid,))
        conn.execute("DELETE FROM templates WHERE campaign_id=?", (cid,))
        conn.execute("DELETE FROM campaigns WHERE id=?", (cid,))
        conn.commit()

    return {"ok": True, "msg": "Campaign 已删除"}


# Gmail label helper for auto-labeling outreach threads
_label_cache = {}  # email -> label_id

def get_or_create_gmail_label(service, label_name="Backlink Outreach"):
    """Get existing label ID or create a new one. Cached per session."""
    cache_key = label_name
    if cache_key in _label_cache:
        return _label_cache[cache_key]
    try:
        results = service.users().labels().list(userId='me').execute()
        for lbl in results.get('labels', []):
            if lbl['name'] == label_name:
                _label_cache[cache_key] = lbl['id']
                return lbl['id']
        # Create label if not found
        label_body = {
            'name': label_name,
            'labelListVisibility': 'labelShow',
            'messageListVisibility': 'show'
        }
        created = service.users().labels().create(userId='me', body=label_body).execute()
        _label_cache[cache_key] = created['id']
        return created['id']
    except Exception as e:
        print(f"[Label] Error: {e}")
        return None

def send_gmail(account_email: str, to: str, subject: str, body: str, cc: str = '',
               thread_id: str = None, in_reply_to: str = None) -> dict:
    """Send email via Gmail API. Returns dict with ok, message_id, thread_id."""
    if TEST_MODE:
        print(f"[TEST MODE] Would send email:")
        print(f"  From: {account_email}")
        print(f"  To: {to}")
        print(f"  Subject: {subject}")
        print(f"  CC: {cc}")
        print(f"  Body: {body[:200]}...")
        return {'ok': True, 'message_id': '', 'thread_id': ''}

    with get_db() as conn:
        row = conn.execute("SELECT token FROM accounts WHERE email=?", (account_email,)).fetchone()
        if not row: return {'ok': False, 'message_id': '', 'thread_id': ''}
    creds = Credentials.from_authorized_user_info(json.loads(row['token']))
    service = build_gmail_service(creds)
    msg = MIMEMultipart('alternative')
    msg['To'], msg['Subject'] = to, subject
    if cc:
        msg['Cc'] = cc
    # ✏️ Threading: set In-Reply-To and References for follow-up emails
    if in_reply_to:
        msg['In-Reply-To'] = in_reply_to
        msg['References'] = in_reply_to
    msg.attach(MIMEText(body, 'html'))
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    # ✏️ Threading: pass threadId so Gmail groups messages in the same thread
    send_body = {'raw': raw}
    if thread_id:
        send_body['threadId'] = thread_id
    sent = service.users().messages().send(userId='me', body=send_body).execute()

    # Extract Message-ID header from sent message for future threading
    sent_message_id = ''
    sent_thread_id = sent.get('threadId', '')
    try:
        sent_detail = service.users().messages().get(
            userId='me', id=sent['id'], format='metadata',
            metadataHeaders=['Message-ID']
        ).execute()
        hdrs = sent_detail.get('payload', {}).get('headers', [])
        sent_message_id = next(
            (h['value'] for h in hdrs if h.get('name', '').lower() == 'message-id'),
            ''
        )
    except Exception as e:
        print(f"[Threading] Could not retrieve Message-ID: {e}")

    # Auto-label the thread for easy tracking
    try:
        label_id = get_or_create_gmail_label(service)
        if label_id and sent_thread_id:
            service.users().threads().modify(
                userId='me', id=sent_thread_id,
                body={'addLabelIds': [label_id]}
            ).execute()
    except Exception as e:
        print(f"[Label] Could not label thread: {e}")
    return {'ok': True, 'message_id': sent_message_id, 'thread_id': sent_thread_id}

def _header_map(message: dict) -> dict:
    headers = message.get('payload', {}).get('headers', [])
    return {h.get('name', '').lower(): h.get('value', '') for h in headers}

def _extract_email(header_value: str) -> str:
    from email.utils import getaddresses
    addresses = getaddresses([header_value or ''])
    for _, addr in addresses:
        if addr:
            return addr.lower().strip()
    match = re.search(r'<(.+?)>', header_value or '')
    return (match.group(1) if match else (header_value or '')).lower().strip()

def _parse_email_time(value: str):
    from email.utils import parsedate_to_datetime
    from dateutil import parser as dateparser
    if not value:
        return None
    try:
        parsed = parsedate_to_datetime(value)
    except Exception:
        parsed = dateparser.parse(value)
    if parsed and parsed.tzinfo is not None:
        parsed = parsed.replace(tzinfo=None)
    return parsed

def _is_auto_reply(sender: str, headers: dict) -> bool:
    local_part = sender.split('@', 1)[0].lower()
    subject = headers.get('subject', '').lower()
    auto_submitted = headers.get('auto-submitted', '').lower()
    precedence = headers.get('precedence', '').lower()
    if local_part in {'mailer-daemon', 'postmaster'}:
        return True
    if local_part in {'no-reply', 'noreply', 'do-not-reply'}:
        return True
    if auto_submitted and auto_submitted != 'no':
        return True
    if precedence in {'bulk', 'junk', 'list'}:
        return True
    if 'delivery status notification' in subject:
        return True
    return False

def _message_is_human_inbound(message: dict, account_email: str) -> tuple[bool, str, dict]:
    headers = _header_map(message)
    sender = _extract_email(headers.get('from', ''))
    labels = set(message.get('labelIds', []))
    if not sender:
        return False, sender, headers
    if sender == account_email.lower().strip() or 'SENT' in labels:
        return False, sender, headers
    if _is_auto_reply(sender, headers):
        return False, sender, headers
    return True, sender, headers

def check_replies(cid: int, account_email: str):
    """检查Gmail thread/inbox，标记已回复的leads"""
    if TEST_MODE:
        print(f"[TEST MODE] Skipping reply check for campaign {cid}")
        return

    try:
        with get_db() as conn:
            row = conn.execute("SELECT token FROM accounts WHERE email=?", (account_email,)).fetchone()
            if not row:
                return

            token_json = row['token']
            leads = conn.execute(
                """
                SELECT id, email, last_sent_at, thread_id
                FROM leads
                WHERE campaign_id=? AND status='pending' AND replied=0 AND last_sent_at IS NOT NULL
                """,
                (cid,)
            ).fetchall()
            if not leads:
                return

        creds = Credentials.from_authorized_user_info(json.loads(token_json))
        service = build_gmail_service(creds)
        replied_lead_ids = []

        for lead in leads:
            lead_id = lead['id']
            lead_email = lead['email'].lower().strip()
            try:
                last_sent_time = _parse_email_time(lead['last_sent_at'])
            except Exception as e:
                print(f"[Date parse error for {lead_email}] {e}")
                continue

            thread_id = lead['thread_id']
            if thread_id:
                try:
                    thread = service.users().threads().get(
                        userId='me', id=thread_id, format='metadata',
                        metadataHeaders=[
                            'From', 'Date', 'Subject', 'Auto-Submitted',
                            'Precedence', 'Message-ID', 'In-Reply-To', 'References'
                        ]
                    ).execute()
                    messages = thread.get('messages', [])
                    sent_times = []
                    inbound_messages = []

                    for message in messages:
                        headers = _header_map(message)
                        sender = _extract_email(headers.get('from', ''))
                        labels = set(message.get('labelIds', []))
                        msg_time = _parse_email_time(headers.get('date', ''))
                        if msg_time and (sender == account_email.lower().strip() or 'SENT' in labels):
                            sent_times.append(msg_time)
                        is_inbound, inbound_sender, inbound_headers = _message_is_human_inbound(message, account_email)
                        if is_inbound:
                            inbound_messages.append((inbound_sender, inbound_headers, msg_time))

                    first_sent_time = min(sent_times) if sent_times else last_sent_time
                    for sender, headers, received_time in inbound_messages:
                        if received_time and first_sent_time and received_time <= first_sent_time:
                            continue
                        replied_lead_ids.append((lead_id, sender))
                        print(f"[Reply] lead {lead_id} thread reply from {sender} (first sent: {first_sent_time}, received: {received_time})")
                        break
                    continue
                except Exception as e:
                    print(f"[Thread reply check error for lead {lead_id}] {e}")

            try:
                results = service.users().messages().list(
                    userId='me', q=f'from:{lead_email} newer_than:30d', maxResults=5
                ).execute()
                for msg in results.get('messages', []):
                    msg_detail = service.users().messages().get(
                        userId='me', id=msg['id'], format='metadata',
                        metadataHeaders=['From', 'Date', 'Subject', 'Auto-Submitted', 'Precedence']
                    ).execute()
                    is_inbound, sender, headers = _message_is_human_inbound(msg_detail, account_email)
                    if not is_inbound or sender != lead_email:
                        continue
                    received_time = _parse_email_time(headers.get('date', ''))
                    if received_time and last_sent_time and received_time <= last_sent_time:
                        continue
                    replied_lead_ids.append((lead_id, sender))
                    print(f"[Reply] lead {lead_id} direct reply from {sender} (sent: {last_sent_time}, received: {received_time})")
                    break
            except Exception as e:
                print(f"[Direct reply check error for lead {lead_id}] {e}")

        if replied_lead_ids:
            seen = set()
            unique_replies = []
            for lead_id, sender in replied_lead_ids:
                if lead_id not in seen:
                    seen.add(lead_id)
                    unique_replies.append((lead_id, sender))

            with get_db() as conn:
                for lead_id, sender in unique_replies:
                    conn.execute("UPDATE leads SET replied=1, opened=1, status='replied' WHERE id=?", (lead_id,))
                    print(f"[Reply detected] Lead {lead_id} ({sender}) replied")
                conn.commit()

    except Exception as e:
        print(f"[Check replies error] {e}")

def _get_campaign_process_lock(cid: int):
    with campaign_process_locks_guard:
        if cid not in campaign_process_locks:
            campaign_process_locks[cid] = threading.Lock()
        return campaign_process_locks[cid]


def process_campaign(cid: int, account_email: str):
    lock = _get_campaign_process_lock(cid)
    if not lock.acquire(blocking=False):
        print(f"[Campaign {cid}] process already running, skip duplicate trigger")
        return
    try:
        _process_campaign_locked(cid, account_email)
    finally:
        lock.release()


def _process_campaign_locked(cid: int, account_email: str):
    with get_db() as conn:
        campaign = conn.execute("SELECT status FROM campaigns WHERE id=?", (cid,)).fetchone()
        if not campaign or campaign['status'] != 'running':
            return

    # 先检查回复
    check_replies(cid, account_email)

    with get_db() as conn:
        lead = conn.execute("""
            SELECT l.* FROM leads l WHERE l.campaign_id=? AND l.status='pending' AND l.replied=0
            AND (l.last_sent_at IS NULL OR datetime(l.last_sent_at, '+' ||
                (SELECT ((COALESCE(delay_days, 0) * 24) + COALESCE(delay_hours, 0))
                 FROM templates WHERE campaign_id=? AND step=l.current_step) || ' hours') <= datetime('now'))
            LIMIT 1
        """, (cid, cid)).fetchone()
        if not lead: return

        tpl = conn.execute("SELECT * FROM templates WHERE campaign_id=? AND step=?",
                          (cid, lead['current_step'])).fetchone()
        if not tpl:
            conn.execute("UPDATE leads SET status='completed' WHERE id=?", (lead['id'],))
            conn.commit(); return

        data = json.loads(lead['data'])
        data['email'] = lead['email']
        subject = render_template_text(tpl['subject'], data)
        body = render_template_text(tpl['body'], data)

        # ✏️ CHANGED: Auto-detect plain text vs HTML — plain text gets newlines converted to <br>
        body = prepare_email_body(body)

        # ✏️ CHANGED: Fixed tracking pixel for accurate open rate tracking
        # - Removed style="display:none" (some email clients skip loading hidden images)
        # - Added ?t= cache-busting parameter (prevents email proxies from caching)
        # - Added alt="" and border="0" for broader email client compatibility
        cache_bust = int(time.time())
        tracker_url = os.environ.get("TRACKER_URL", "")
        if tracker_url:
            track_pixel = f'<img src="{tracker_url}/open?uid={lead["id"]}&t={cache_bust}" width="1" height="1" alt="" border="0">'
        else:
            base_url = get_tracking_base_url()
            track_pixel = f'<img src="{base_url}/track/open/{lead["id"]}?t={cache_bust}" width="1" height="1" alt="" border="0">'
        body += track_pixel

        # ✏️ CC support: read cc_email from campaign settings
        campaign_row = conn.execute("SELECT cc_email FROM campaigns WHERE id=?", (cid,)).fetchone()
        cc_email = campaign_row['cc_email'] if campaign_row and campaign_row['cc_email'] else ''

        # ✏️ Threading: pass stored thread_id/message_id for follow-up steps
        _thread_id = lead['thread_id'] if lead['thread_id'] else None
        _in_reply_to = lead['message_id'] if lead['message_id'] else None
        if lead['current_step'] > 1:
            subject = get_followup_subject(conn, cid, data, subject)

        result = send_gmail(account_email, lead['email'], subject, body, cc=cc_email,
                           thread_id=_thread_id, in_reply_to=_in_reply_to)
        if result['ok']:
            next_tpl = conn.execute("SELECT * FROM templates WHERE campaign_id=? AND step=?",
                                   (cid, lead['current_step'] + 1)).fetchone()
            new_status = 'pending' if next_tpl else 'completed'
            # Store threading info so future steps land in the same thread
            new_msg_id = result.get('message_id', '') or lead['message_id'] or ''
            new_thread_id = result.get('thread_id', '') or lead['thread_id'] or ''
            conn.execute(
                "UPDATE leads SET current_step=current_step+1, last_sent_at=?, status=?, message_id=?, thread_id=? WHERE id=?",
                (datetime.now().isoformat(), new_status, new_msg_id, new_thread_id, lead['id']))
            conn.commit()

# 追踪
@app.get("/track/open/{lead_id}")
def track_open(lead_id: int, request: Request):
    """追踪邮件打开"""
    # 记录日志（包含IP和User-Agent）
    client_ip = request.client.host if request.client else "unknown"
    user_agent = request.headers.get("user-agent", "unknown")
    print(f"[Open tracked] Lead {lead_id} | IP: {client_ip} | UA: {user_agent[:50]}...")

    with get_db() as conn:
        # 检查lead是否存在
        lead = conn.execute("SELECT id, email FROM leads WHERE id=?", (lead_id,)).fetchone()
        if lead:
            conn.execute("UPDATE leads SET opened=1 WHERE id=?", (lead_id,))
            conn.commit()
            print(f"[Open tracked] ✓ Lead {lead_id} ({lead['email']}) marked as opened")
        else:
            print(f"[Open tracked] ✗ Lead {lead_id} not found")

    # ✏️ CHANGED: Use Response (not HTMLResponse) with no-cache headers
    # This ensures email clients/proxies don't cache the pixel and miss future opens
    gif = base64.b64decode("R0lGODlhAQABAIAAAAAAAP///yH5BAEAAAAALAAAAAABAAEAAAIBRAA7")
    return Response(
        content=gif,
        media_type="image/gif",
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        }
    )

@app.get("/track/click/{lead_id}")
def track_click(lead_id: int, url: str):
    with get_db() as conn:
        conn.execute("UPDATE leads SET clicked=1 WHERE id=?", (lead_id,))
        conn.commit()
    return RedirectResponse(url)

@app.get("/api/accounts")
def list_accounts():
    with get_db() as conn:
        rows = conn.execute("SELECT id, email, created_at FROM accounts").fetchall()
        accounts = [dict(r) for r in rows]
        # 测试模式下添加虚拟账号
        if TEST_MODE and not accounts:
            conn.execute("INSERT OR IGNORE INTO accounts(email, token) VALUES(?, ?)",
                        ("test@example.com", "{}"))
            conn.commit()
            accounts = [{"id": 1, "email": "test@example.com", "created_at": "test"}]
        return accounts

@app.get("/api/campaigns/{cid}/stats")
def campaign_stats(cid: int):
    with get_db() as conn:
        stats = conn.execute("""
            SELECT COUNT(*) as total, SUM(opened) as opens, SUM(clicked) as clicks, SUM(replied) as replies,
            SUM(CASE WHEN status IN ('completed', 'replied') OR current_step > 1 THEN 1 ELSE 0 END) as sent
            FROM leads WHERE campaign_id=?
        """, (cid,)).fetchone()
        return dict(stats)

@app.post("/api/campaigns/{cid}/check-replies")
def check_replies_now(cid: int):
    """手动触发检查回复"""
    with get_db() as conn:
        campaign = conn.execute("SELECT account_email FROM campaigns WHERE id=?", (cid,)).fetchone()
        if not campaign or not campaign['account_email']:
            return {"msg": "请先启动campaign以设置发件账号"}
    check_replies(cid, campaign['account_email'])
    return {"msg": "检查完成"}

@app.post("/api/leads/{lead_id}/mark")
def mark_lead(lead_id: int, field: str):
    """手动标记lead状态（用于测试）"""
    if field not in ('opened', 'clicked', 'replied'):
        raise HTTPException(400, "Invalid field")
    with get_db() as conn:
        if field == 'replied':
            conn.execute("UPDATE leads SET replied=1, opened=1, status='replied' WHERE id=?", (lead_id,))
        else:
            conn.execute(f"UPDATE leads SET {field}=1 WHERE id=?", (lead_id,))
        conn.commit()
    return {"ok": True}

# ✏️ CHANGED: New settings API endpoints for configuring BASE_URL from the web UI
@app.get("/api/settings")
def get_settings():
    """Get all settings"""
    with get_db() as conn:
        rows = conn.execute("SELECT key, value FROM settings").fetchall()
        settings = {r['key']: r['value'] for r in rows}
    # Include current effective base_url even if not saved yet
    if 'base_url' not in settings:
        settings['base_url'] = os.environ.get("BASE_URL", "http://localhost:8000")
    return settings

@app.post("/api/settings")
async def update_settings(request: Request):
    """Update settings (e.g. base_url)"""
    data = await request.json()
    with get_db() as conn:
        for key, value in data.items():
            conn.execute("INSERT OR REPLACE INTO settings(key, value) VALUES(?, ?)", (key, value))
        conn.commit()
    return {"ok": True}

# ============================================
# 自动同步追踪数据（从Render tracker拉取）
# ============================================

def sync_tracker_data():
    """
    从Render追踪服务同步数据到本地数据库
    每10分钟自动运行一次
    """
    import requests

    tracker_url = os.environ.get("TRACKER_URL", "")

    if not tracker_url or "your-tracker" in tracker_url:
        # 未配置追踪URL，跳过同步
        return

    try:
        # 获取未同步的打开记录
        response = requests.get(f"{tracker_url}/api/opens", timeout=10)
        if response.status_code != 200:
            print(f"[Sync Warning] Tracker API returned {response.status_code}")
            return

        data = response.json()
        opens = data.get("opens", [])

        if not opens:
            return

        # 更新本地数据库
        with get_db() as conn:
            updated = 0
            synced_ids = []

            for record in opens:
                uid = record['uid']
                record_id = record['id']

                try:
                    # 检查lead是否存在且未标记打开
                    lead = conn.execute(
                        "SELECT opened FROM leads WHERE id=?",
                        (uid,)
                    ).fetchone()

                    if lead and lead[0] == 0:
                        conn.execute("UPDATE leads SET opened=1 WHERE id=?", (uid,))
                        updated += 1

                    synced_ids.append(record_id)

                except Exception as e:
                    print(f"[Sync Error] Lead {uid}: {e}")

            conn.commit()

        # 标记远程记录为已同步
        if synced_ids:
            try:
                requests.post(
                    f"{tracker_url}/api/mark_synced",
                    json={"open_ids": synced_ids, "click_ids": []},
                    timeout=5
                )
                print(f"[Sync] Updated {updated} opens, marked {len(synced_ids)} as synced")
            except:
                pass  # 忽略标记失败

        # 同步点击记录
        response = requests.get(f"{tracker_url}/api/clicks", timeout=10)
        if response.status_code == 200:
            clicks = response.json().get("clicks", [])
            if clicks:
                with get_db() as conn:
                    click_ids = []
                    for record in clicks:
                        try:
                            conn.execute("UPDATE leads SET clicked=1 WHERE id=?", (record['uid'],))
                            click_ids.append(record['id'])
                        except:
                            pass
                    conn.commit()

                if click_ids:
                    requests.post(
                        f"{tracker_url}/api/mark_synced",
                        json={"open_ids": [], "click_ids": click_ids},
                        timeout=5
                    )

    except requests.exceptions.RequestException as e:
        print(f"[Sync Error] {e}")
    except Exception as e:
        print(f"[Sync Error] Unexpected: {e}")

# 添加自动同步任务（每10分钟）
if os.environ.get("TRACKER_URL"):
    scheduler.add_job(
        sync_tracker_data,
        'interval',
        minutes=10,
        id='sync_tracker',
        replace_existing=True
    )
    print("[Scheduler] Tracker sync enabled (every 10 minutes)")

# 前端页面
@app.get("/", response_class=HTMLResponse)
def index():
    content = Path("index.html").read_text(encoding='utf-8')
    return HTMLResponse(content=content, headers={
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
    })

# Serve static files (logos, images)
app.mount("/static", StaticFiles(directory="static"), name="static")

# 启动时恢复运行中的 campaigns（延迟执行，确保所有函数已定义）
scheduler.add_job(restore_running_campaigns, 'date', id='restore_on_start')

if __name__ == "__main__":
    # ✏️ CHANGED: Startup warning when BASE_URL is localhost (tracking won't work)
    effective_url = get_tracking_base_url()
    if "localhost" in effective_url or "127.0.0.1" in effective_url:
        print("\n" + "=" * 60)
        print("  WARNING: BASE_URL is set to localhost.")
        print("  Email open tracking will NOT work!")
        print("  ")
        print("  To fix: open the web UI and set a public URL in Settings,")
        print("  or use ngrok:  ngrok http 8000")
        print("=" * 60 + "\n")

    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
